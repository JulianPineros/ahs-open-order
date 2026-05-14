import streamlit as st
import io
import re
import datetime

st.set_page_config(page_title="AHS ↔ Odoo | Fechas de carga", page_icon="📦", layout="centered")

st.title("📦 AHS — Fechas de carga estimada")
st.markdown("Sube los dos archivos y descarga el Excel con las fechas proyectadas.")

# ── Dependencias internas ─────────────────────────────────────────────────────
try:
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("Faltan dependencias. Verifica requirements.txt")
    st.stop()

DIAS_LLEGADA = 70

# ── Helpers ───────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center"):
    return Alignment(horizontal=h, vertical="center")

def _mdy_to_dmy(date_str):
    try:
        m, d, y = date_str.strip().split("/")
        return f"{int(d):02d}/{int(m):02d}/{y}"
    except Exception:
        return date_str

def _parse_dmy(date_str):
    try:
        d, m, y = date_str.strip().split("/")
        return datetime.date(int(y), int(m), int(d))
    except Exception:
        return None

def _earlier(a, b):
    da, db = _parse_dmy(a), _parse_dmy(b)
    if da and db:
        return a if da <= db else b
    return a or b

def _decode(file_bytes):
    if file_bytes[:2] in (b'\xff\xfe', b'\xfe\xff'):
        return file_bytes.decode("utf-16")
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return file_bytes.decode("latin-1", errors="replace")

# ── Parser Open Order (formato HTML con encabezados por pedido) ───────────────
# Cruza por (N.º de pedido, Artículo n.º) → Fecha carga est.
def parse_open_order(file_bytes):
    content = _decode(file_bytes)
    soup = BeautifulSoup(content, "lxml")

    pedidos     = set()
    idx_ref_sku = {}   # (pedido, sku) → fecha DD/MM/YYYY
    idx_sku     = {}   # sku → fecha DD/MM/YYYY  (fallback sin referencia)
    current_order = {}

    for table in soup.find_all("table"):
        cells_per_row = [
            [td.get_text(strip=True) for td in row.find_all("td")]
            for row in table.find_all("tr")
        ]
        flat = " | ".join(c for row in cells_per_row for c in row)

        # Tabla de encabezado de pedido
        if "N.º de tienda:" in flat and "N.º de pedido:" in flat and "Artículo n.º" not in flat:
            m = re.search(r"N\.º de pedido:\s*\|\s*(\S+)", flat)
            current_order = {"order": m.group(1) if m else ""}
            if current_order["order"]:
                pedidos.add(current_order["order"])

        # Tabla de artículos
        elif "Artículo n.º" in flat and current_order:
            for rc in cells_per_row:
                if not rc or any("Artículo n.º" in c for c in rc) or "TOTAL" in (rc[0] if rc else ""):
                    continue
                if len(rc) >= 15 and len(rc[2]) == 1 and rc[2].isalpha():
                    sku    = rc[1].strip()
                    fecha  = _mdy_to_dmy(rc[14].strip())
                    pedido = current_order.get("order", "")
                    key    = (pedido, sku)
                    idx_ref_sku[key] = _earlier(idx_ref_sku[key], fecha) if key in idx_ref_sku else fecha
                    idx_sku[sku]     = _earlier(idx_sku[sku], fecha)     if sku in idx_sku     else fecha

    return pedidos, idx_ref_sku, idx_sku

# ── Parser Trasladar (Odoo stock.picking) ────────────────────────────────────
def parse_trasladar(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    headers  = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows     = []
    last_ref = ""
    ref_known = False

    for r in range(2, ws.max_row + 1):
        id_col  = ws.cell(r, 1).value
        ref_col = ws.cell(r, 2).value
        sku     = ws.cell(r, 5).value
        vals    = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

        if id_col and str(id_col).strip():
            if ref_col and str(ref_col).strip():
                last_ref  = str(ref_col).strip()
                ref_known = True
            else:
                last_ref  = ""
                ref_known = False

        rows.append({
            "vals":        vals,
            "_ref_filled": last_ref,
            "_ref_known":  ref_known,
            "_sku":        str(sku).strip() if sku else "",
        })

    return headers, rows

# ── Generar Excel ─────────────────────────────────────────────────────────────
def generate_excel(headers, rows, idx_ref_sku, idx_sku):
    HEADER_FILL = "1F4E79"
    HEADER_NEW  = "2E75B6"
    MATCH_FILL  = ("E2EFDA", "EBF5E1")
    FALLBK_FILL = ("FFF2CC", "FFFACD")
    TRANS_FILL  = ("FCE4D6", "FDE9E7")

    NEW_COLS = ["Fecha carga est. (Ashley)", f"Fecha est. llegada (+{DIAS_LLEGADA}d)"]
    all_cols = headers + NEW_COLS
    n_orig   = len(headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Odoo + Fechas Ashley"

    for c, name in enumerate(all_cols, 1):
        cell = ws.cell(1, c, name)
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _fill(HEADER_FILL if c <= n_orig else HEADER_NEW)
        cell.alignment = _align()
        cell.border    = _border()
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"

    for c, w in enumerate([40, 32, 45, 20, 18, 10, 24, 24], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    stats = {"exact": 0, "fallback": 0, "transit": 0}

    for i, row in enumerate(rows):
        r_idx     = i + 2
        vals      = row["vals"]
        ref       = row["_ref_filled"]
        ref_known = row["_ref_known"]
        sku       = row["_sku"]

        fecha_carga = ""
        fecha_llega = ""
        match_type  = "transit"

        # 1. Match exacto: pedido + SKU
        if ref and sku:
            result = idx_ref_sku.get((ref, sku))
            if result:
                fecha_carga = result
                match_type  = "exact"

        # 2. Fallback por SKU solo (cuando el picking no tiene referencia conocida)
        if not fecha_carga and not ref_known and sku:
            result = idx_sku.get(sku)
            if result:
                fecha_carga = result
                match_type  = "fallback"

        if fecha_carga:
            d = _parse_dmy(fecha_carga)
            if d:
                fecha_llega = (d + datetime.timedelta(days=DIAS_LLEGADA)).strftime("%d/%m/%Y")

        stats[match_type] += 1
        palette  = {"exact": MATCH_FILL, "fallback": FALLBK_FILL, "transit": TRANS_FILL}
        row_fill = palette[match_type][i % 2]

        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.fill      = _fill(row_fill)
            cell.border    = _border()
            cell.alignment = _align("left" if c_idx in (1, 2, 3) else "center")
            cell.font      = _font()

        for c_idx, val in enumerate([fecha_carga, fecha_llega], n_orig + 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.fill      = _fill(MATCH_FILL[i % 2] if val else TRANS_FILL[i % 2])
            cell.border    = _border()
            cell.alignment = _align()
            cell.font      = _font(bold=bool(val))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), stats

# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
col1, col2 = st.columns(2)

with col1:
    st.subheader("1️⃣ Open Order Report")
    st.caption("Archivo .xls exportado de Ashley")
    oo_file = st.file_uploader("", type=["xls"], key="oo")

with col2:
    st.subheader("2️⃣ Trasladar (Odoo)")
    st.caption("Archivo .xlsx exportado de Odoo (stock.picking)")
    tr_file = st.file_uploader("", type=["xlsx"], key="tr")

st.divider()

if oo_file and tr_file:
    if st.button("⚙️ Procesar archivos", type="primary", use_container_width=True):
        with st.spinner("Procesando..."):
            try:
                oo_bytes = oo_file.read()
                tr_bytes = tr_file.read()

                pedidos, idx_ref_sku, idx_sku = parse_open_order(oo_bytes)
                headers, rows_all = parse_trasladar(tr_bytes)

                rows = [r for r in rows_all if r["_ref_filled"] in pedidos]

                ts          = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                output_name = f"Odoo_FechasAshley_{ts}.xlsx"
                excel_bytes, stats = generate_excel(headers, rows, idx_ref_sku, idx_sku)

                st.success("✅ Proceso completado")

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Pedidos Ashley", len(pedidos))
                m2.metric("Filas procesadas", len(rows))
                m3.metric("Match exacto", stats["exact"])
                m4.metric("Sin fecha", stats["transit"])

                st.info(
                    f"🟢 **{stats['exact']} filas** con match exacto (pedido + SKU)  \n"
                    f"🟡 **{stats['fallback']} filas** con match por SKU solo  \n"
                    f"🟠 **{stats['transit']} filas** sin fecha en Ashley"
                )

                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_bytes,
                    file_name=output_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )

            except Exception as e:
                st.error(f"Error al procesar: {e}")
                st.exception(e)
else:
    st.info("👆 Sube los dos archivos para continuar.")

st.divider()
st.caption("AHS – Nirvana  ·  Proceso automatizado de fechas de carga estimada")
